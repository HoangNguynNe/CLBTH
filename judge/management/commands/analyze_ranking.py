"""
Django management command để phân tích sao chép code và xuất bảng xếp hạng ra Excel
Sử dụng: python manage.py analyze_ranking
"""

import os
import re
import hashlib
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User

from judge.models import Submission, SubmissionSource, Profile


class Command(BaseCommand):
    help = "Phân tích sao chép code và xuất bảng xếp hạng ra Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "-o",
            "--output",
            default="ranking_report.xlsx",
            help="Tên file Excel đầu ra",
        )

    def normalize_code_exact(self, code):
        """Chuẩn hóa code để so sánh exact copy - chỉ bỏ khoảng trắng thừa"""
        if not code:
            return ""
        # Bỏ comment
        code = re.sub(r"//.*?$", "", code, flags=re.MULTILINE)
        code = re.sub(r"/\*.*?\*/", "", code, flags=re.DOTALL)
        code = re.sub(r"#.*?$", "", code, flags=re.MULTILINE)  # Python comments
        # Bỏ khoảng trắng thừa
        code = re.sub(r"\s+", " ", code).strip()
        return code

    def get_code_hash(self, code):
        """Tạo hash của code"""
        return hashlib.md5(code.encode("utf-8", errors="ignore")).hexdigest()

    def extract_variable_names(self, code):
        """Trích xuất tên biến từ code"""
        if not code:
            return set()

        # Bỏ comments và strings
        code = re.sub(r"//.*?$", "", code, flags=re.MULTILINE)
        code = re.sub(r"/\*.*?\*/", "", code, flags=re.DOTALL)
        code = re.sub(r"#.*?$", "", code, flags=re.MULTILINE)
        code = re.sub(r'"[^"]*"', "", code)
        code = re.sub(r"'[^']*'", "", code)

        # Tìm các identifier (tên biến, hàm)
        identifiers = re.findall(r"\b[a-zA-Z_][a-zA-Z0-9_]*\b", code)

        # Loại bỏ các keyword
        keywords = {
            "if",
            "else",
            "for",
            "while",
            "do",
            "switch",
            "case",
            "break",
            "continue",
            "return",
            "int",
            "long",
            "float",
            "double",
            "char",
            "string",
            "bool",
            "void",
            "unsigned",
            "signed",
            "short",
            "auto",
            "class",
            "struct",
            "public",
            "private",
            "protected",
            "namespace",
            "using",
            "include",
            "define",
            "ifdef",
            "ifndef",
            "endif",
            "pragma",
            "const",
            "static",
            "extern",
            "virtual",
            "override",
            "new",
            "delete",
            "true",
            "false",
            "null",
            "nullptr",
            "this",
            "template",
            "typename",
            "try",
            "catch",
            "throw",
            "sizeof",
            "typedef",
            "enum",
            "union",
            "main",
            "cin",
            "cout",
            "endl",
            "scanf",
            "printf",
            "gets",
            "puts",
            "std",
            "vector",
            "map",
            "set",
            "pair",
            "queue",
            "stack",
            "deque",
            "sort",
            "min",
            "max",
            "abs",
            "swap",
            "push_back",
            "pop_back",
            "begin",
            "end",
            "size",
            "empty",
            "clear",
            "insert",
            "erase",
            "find",
            "def",
            "print",
            "input",
            "range",
            "len",
            "str",
            "list",
            "dict",
            "import",
            "from",
            "as",
            "in",
            "and",
            "or",
            "not",
            "is",
            "lambda",
            "pass",
            "with",
            "assert",
            "yield",
            "global",
            "nonlocal",
            "raise",
            "except",
            "finally",
            "True",
            "False",
            "None",
            "self",
            "cls",
        }

        return set(identifiers) - keywords

    def get_structure_fingerprint(self, code):
        """
        Tạo fingerprint dựa trên cấu trúc code
        Dùng để phát hiện form copy kể cả khi đổi tên biến
        """
        if not code:
            return ""

        # Bỏ comments
        code = re.sub(r"//.*?$", "", code, flags=re.MULTILINE)
        code = re.sub(r"/\*.*?\*/", "", code, flags=re.DOTALL)
        code = re.sub(r"#.*?$", "", code, flags=re.MULTILINE)

        # Thay thế string literals
        code = re.sub(r'"[^"]*"', "S", code)
        code = re.sub(r"'[^']*'", "S", code)

        # Thay thế số
        code = re.sub(r"\b\d+\.?\d*\b", "N", code)

        # Thay thế tất cả identifier thành X
        code = re.sub(r"\b[a-zA-Z_][a-zA-Z0-9_]*\b", "X", code)

        # Giữ lại cấu trúc (dấu ngoặc, toán tử, etc.)
        code = re.sub(r"\s+", "", code)

        return code

    def analyze_submissions(self):
        """Phân tích tất cả submissions"""
        self.stdout.write("Đang lấy dữ liệu submissions...")

        # Lấy tất cả submissions có source code, sắp xếp theo thời gian
        submissions = (
            Submission.objects.select_related("user", "user__user", "problem")
            .prefetch_related("source")
            .filter(status="D")  # Chỉ lấy completed submissions
            .order_by("date")
        )

        total = submissions.count()
        self.stdout.write(f"Tổng số submissions: {total}")

        # Dictionary lưu trữ code theo problem
        problem_codes = defaultdict(list)

        # User data
        user_data = {}

        count = 0
        for sub in submissions:
            count += 1
            if count % 500 == 0:
                self.stdout.write(f"Đang xử lý: {count}/{total}")

            try:
                source = sub.source.source
            except:
                continue

            user_id = sub.user_id
            if user_id not in user_data:
                user = sub.user.user
                user_data[user_id] = {
                    "username": user.username,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "submissions": [],
                    "problems_attempted": set(),
                    "problems_ac": set(),
                    "problems_wa": set(),
                    "problems_copied_exact": set(),
                    "problems_copied_form": set(),
                    "problems_valid_after_exact": set(),
                    "problems_valid_after_form": set(),
                }

            user_data[user_id]["submissions"].append(
                {
                    "id": sub.id,
                    "problem_id": sub.problem_id,
                    "problem_code": sub.problem.code,
                    "date": sub.date,
                    "result": sub.result,
                    "source": source,
                }
            )

            user_data[user_id]["problems_attempted"].add(sub.problem_id)
            if sub.result == "AC":
                user_data[user_id]["problems_ac"].add(sub.problem_id)
            elif sub.result == "WA":
                user_data[user_id]["problems_wa"].add(sub.problem_id)

            problem_codes[sub.problem_id].append(
                {
                    "submission_id": sub.id,
                    "user_id": user_id,
                    "date": sub.date,
                    "source": source,
                    "result": sub.result,
                }
            )

        self.stdout.write("\nĐang phát hiện sao chép...")

        # Phát hiện exact copy
        exact_copied = defaultdict(set)
        form_copied = defaultdict(set)

        # Chi tiết copy: [(copier_user_id, copier_username, original_user_id, original_username,
        #                  problem_code, copier_submission_id, original_submission_id, copy_type)]
        copy_details = []

        # Lưu problem_code mapping
        problem_code_map = {}

        for problem_id, subs in problem_codes.items():
            if len(subs) < 2:
                continue

            # Lấy problem code từ submission đầu tiên
            if problem_id not in problem_code_map:
                first_sub = subs[0]
                # Tìm problem code từ user_data
                for uid, udata in user_data.items():
                    for s in udata["submissions"]:
                        if s["problem_id"] == problem_id:
                            problem_code_map[problem_id] = s["problem_code"]
                            break
                    if problem_id in problem_code_map:
                        break

            # Sắp xếp theo thời gian
            subs.sort(key=lambda x: x["date"])

            # So sánh exact copy
            exact_hashes = {}  # {hash: (user_id, submission_id)}
            for sub in subs:
                normalized = self.normalize_code_exact(sub["source"])
                code_hash = self.get_code_hash(normalized)

                if code_hash in exact_hashes:
                    first_user_id, first_sub_id = exact_hashes[code_hash]
                    if first_user_id != sub["user_id"]:
                        exact_copied[sub["user_id"]].add(problem_id)
                        # Lưu chi tiết
                        copy_details.append(
                            {
                                "copier_user_id": sub["user_id"],
                                "copier_username": user_data[sub["user_id"]][
                                    "username"
                                ],
                                "copier_fullname": f"{user_data[sub['user_id']]['last_name']} {user_data[sub['user_id']]['first_name']}",
                                "original_user_id": first_user_id,
                                "original_username": user_data[first_user_id][
                                    "username"
                                ],
                                "original_fullname": f"{user_data[first_user_id]['last_name']} {user_data[first_user_id]['first_name']}",
                                "problem_code": problem_code_map.get(
                                    problem_id, str(problem_id)
                                ),
                                "copier_submission_id": sub["submission_id"],
                                "original_submission_id": first_sub_id,
                                "copy_type": "Exact Copy",
                                "copier_date": sub["date"],
                            }
                        )
                else:
                    exact_hashes[code_hash] = (sub["user_id"], sub["submission_id"])

            # So sánh form copy (cấu trúc + tên biến giống nhau)
            form_fingerprints = {}  # {key: (user_id, submission_id)}
            for sub in subs:
                fingerprint = self.get_structure_fingerprint(sub["source"])
                var_names = self.extract_variable_names(sub["source"])
                var_key = frozenset(var_names) if len(var_names) > 0 else frozenset()

                key = (fingerprint, var_key)

                if key in form_fingerprints:
                    first_user_id, first_sub_id = form_fingerprints[key]
                    if first_user_id != sub["user_id"]:
                        form_copied[sub["user_id"]].add(problem_id)
                        # Lưu chi tiết (chỉ nếu chưa có trong exact copy)
                        is_exact = any(
                            d["copier_submission_id"] == sub["submission_id"]
                            and d["copy_type"] == "Exact Copy"
                            for d in copy_details
                        )
                        if not is_exact:
                            copy_details.append(
                                {
                                    "copier_user_id": sub["user_id"],
                                    "copier_username": user_data[sub["user_id"]][
                                        "username"
                                    ],
                                    "copier_fullname": f"{user_data[sub['user_id']]['last_name']} {user_data[sub['user_id']]['first_name']}",
                                    "original_user_id": first_user_id,
                                    "original_username": user_data[first_user_id][
                                        "username"
                                    ],
                                    "original_fullname": f"{user_data[first_user_id]['last_name']} {user_data[first_user_id]['first_name']}",
                                    "problem_code": problem_code_map.get(
                                        problem_id, str(problem_id)
                                    ),
                                    "copier_submission_id": sub["submission_id"],
                                    "original_submission_id": first_sub_id,
                                    "copy_type": "Form Copy",
                                    "copier_date": sub["date"],
                                }
                            )
                else:
                    form_fingerprints[key] = (sub["user_id"], sub["submission_id"])

        # Cập nhật user_data
        for user_id in user_data:
            user_data[user_id]["problems_copied_exact"] = exact_copied.get(
                user_id, set()
            )
            user_data[user_id]["problems_copied_form"] = form_copied.get(user_id, set())

            ac_problems = user_data[user_id]["problems_ac"]
            user_data[user_id][
                "problems_valid_after_exact"
            ] = ac_problems - exact_copied.get(user_id, set())
            user_data[user_id][
                "problems_valid_after_form"
            ] = ac_problems - form_copied.get(user_id, set())

        # Sắp xếp copy_details theo thời gian
        copy_details.sort(key=lambda x: x["copier_date"])

        return user_data, copy_details

    def create_excel(self, user_data, copy_details, output_file):
        """Tạo file Excel với 4 trang (thêm trang chi tiết copy)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            import subprocess
            import sys

            self.stdout.write("Đang cài đặt openpyxl...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        self.stdout.write(f"\nĐang tạo file Excel: {output_file}")

        wb = Workbook()

        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "STT",
            "Username",
            "Họ",
            "Tên",
            "Tổng bài đã làm",
            "Tổng bài AC",
            "Tổng bài WA",
            "Số bài sao chép",
            "Số bài hợp lệ",
            "Điểm (AC hợp lệ)",
        ]

        def write_sheet(ws, title, user_list, copy_type="none"):
            ws.title = title

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Tính toán và sắp xếp
            ranking_data = []
            for user_id, data in user_list.items():
                total_attempted = len(data["problems_attempted"])
                total_ac = len(data["problems_ac"])
                total_wa = len(data["problems_wa"])

                if copy_type == "none":
                    copied = 0
                    valid = total_ac
                elif copy_type == "exact":
                    copied = len(data["problems_copied_exact"] & data["problems_ac"])
                    valid = len(data["problems_valid_after_exact"])
                else:  # form
                    copied = len(data["problems_copied_form"] & data["problems_ac"])
                    valid = len(data["problems_valid_after_form"])

                ranking_data.append(
                    {
                        "user_id": user_id,
                        "username": data["username"],
                        "first_name": data["first_name"],
                        "last_name": data["last_name"],
                        "total_attempted": total_attempted,
                        "total_ac": total_ac,
                        "total_wa": total_wa,
                        "copied": copied,
                        "valid": valid,
                    }
                )

            # Sắp xếp theo điểm valid giảm dần
            ranking_data.sort(key=lambda x: -x["valid"])

            # Write data
            for row, data in enumerate(ranking_data, 2):
                ws.cell(row=row, column=1, value=row - 1).border = thin_border
                ws.cell(row=row, column=2, value=data["username"]).border = thin_border
                ws.cell(row=row, column=3, value=data["last_name"]).border = thin_border
                ws.cell(
                    row=row, column=4, value=data["first_name"]
                ).border = thin_border
                ws.cell(
                    row=row, column=5, value=data["total_attempted"]
                ).border = thin_border
                ws.cell(row=row, column=6, value=data["total_ac"]).border = thin_border
                ws.cell(row=row, column=7, value=data["total_wa"]).border = thin_border
                ws.cell(row=row, column=8, value=data["copied"]).border = thin_border
                ws.cell(row=row, column=9, value=data["valid"]).border = thin_border
                ws.cell(row=row, column=10, value=data["valid"]).border = thin_border

            # Auto width
            for col in range(1, len(headers) + 1):
                if col <= 26:
                    ws.column_dimensions[chr(64 + col)].width = 15
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20

        def write_copy_details_sheet(ws, title, details):
            """Ghi sheet chi tiết các trường hợp copy"""
            ws.title = title

            detail_headers = [
                "STT",
                "Loại Copy",
                "Mã bài",
                "Người copy (Username)",
                "Người copy (Họ tên)",
                "ID bài nộp copy",
                "Người gốc (Username)",
                "Người gốc (Họ tên)",
                "ID bài nộp gốc",
            ]

            # Write headers
            for col, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row, detail in enumerate(details, 2):
                ws.cell(row=row, column=1, value=row - 1).border = thin_border
                ws.cell(
                    row=row, column=2, value=detail["copy_type"]
                ).border = thin_border
                ws.cell(
                    row=row, column=3, value=detail["problem_code"]
                ).border = thin_border
                ws.cell(
                    row=row, column=4, value=detail["copier_username"]
                ).border = thin_border
                ws.cell(
                    row=row, column=5, value=detail["copier_fullname"]
                ).border = thin_border
                ws.cell(
                    row=row, column=6, value=detail["copier_submission_id"]
                ).border = thin_border
                ws.cell(
                    row=row, column=7, value=detail["original_username"]
                ).border = thin_border
                ws.cell(
                    row=row, column=8, value=detail["original_fullname"]
                ).border = thin_border
                ws.cell(
                    row=row, column=9, value=detail["original_submission_id"]
                ).border = thin_border

            # Auto width
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 22
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["F"].width = 18
            ws.column_dimensions["G"].width = 22
            ws.column_dimensions["H"].width = 25
            ws.column_dimensions["I"].width = 18

        # Sheet 1: Chưa lọc
        ws1 = wb.active
        write_sheet(ws1, "Chưa lọc copy", user_data, "none")

        # Sheet 2: Lọc exact copy
        ws2 = wb.create_sheet()
        write_sheet(ws2, "Lọc copy giống hệt", user_data, "exact")

        # Sheet 3: Lọc form copy
        ws3 = wb.create_sheet()
        write_sheet(ws3, "Lọc copy form bài", user_data, "form")

        # Sheet 4: Chi tiết copy
        ws4 = wb.create_sheet()
        write_copy_details_sheet(ws4, "Chi tiết sao chép", copy_details)

        wb.save(output_file)
        self.stdout.write(self.style.SUCCESS(f"Đã lưu file: {output_file}"))
        self.stdout.write(f"Tổng số trường hợp sao chép phát hiện: {len(copy_details)}")
        return output_file

    def handle(self, *args, **options):
        output_file = options["output"]

        self.stdout.write("=" * 60)
        self.stdout.write("PHÂN TÍCH SAO CHÉP CODE VÀ BẢNG XẾP HẠNG")
        self.stdout.write("=" * 60)

        user_data, copy_details = self.analyze_submissions()

        self.stdout.write(f"\nTổng số người dùng: {len(user_data)}")
        self.stdout.write(f"Tổng số trường hợp sao chép: {len(copy_details)}")

        self.create_excel(user_data, copy_details, output_file)

        self.stdout.write("\n" + "=" * 60)
        self.stdout.write(self.style.SUCCESS("HOÀN TẤT!"))
        self.stdout.write(f"File Excel đã được tạo: {output_file}")
        self.stdout.write("=" * 60)
